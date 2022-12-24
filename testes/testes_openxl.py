from openpyxl import workbook, load_workbook                    #Importação do modulo de edição de excel

x = load_workbook('orcamentacao_auto.xlsx')                     #
xs = x.active
print(xs)
print(xs['A1'].value)                                           #Lê os valores da célula

xs['A1'].value = 'TESTE 2'                                      #Cria valor na célula
x.save('orcamentacao_auto.xlsx')

print(x.sheetnames)                                             #Mostra o nome das folhas existentes

x.create_sheet('Orcamento')                                     #Cria uma nova folha
